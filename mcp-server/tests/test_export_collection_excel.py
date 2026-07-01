"""Tests for tools.export_collection_excel."""

from unittest.mock import patch, MagicMock
import pytest


def _ctx(role="developer"):
    from context import ToolContext
    return ToolContext(session_id="s1", user_id="u1", role=role)


def _fake_db_seq(results):
    """A get_db whose cursor.fetchone/fetchall pop from a queued list per execute."""
    calls = {"i": 0}
    cur = MagicMock()
    def execute(sql, params=None):
        cur._sql = sql
    def fetchone():
        r = results[calls["i"]]; calls["i"] += 1
        return r
    def fetchall():
        r = results[calls["i"]]; calls["i"] += 1
        return r
    cur.execute.side_effect = execute
    cur.fetchone.side_effect = fetchone
    cur.fetchall.side_effect = fetchall
    conn = MagicMock(); conn.cursor.return_value = cur
    from contextlib import contextmanager
    @contextmanager
    def _get():
        yield conn
    return _get


def test_export_writes_xlsx(tmp_path):
    # call order: _menu_and_fields(fetchone) -> dynamic_data(fetchall) -> _workspace_for_session(fetchone)
    results = [
        ('page-ic', '巡检用例', ['admin', 'developer'],
         [{'fieldName': 'caseName', 'label': '用例名称'}, {'fieldName': 'priority', 'label': '优先级'}]),
        [[{'caseName': 'CPU', 'priority': '高'}], [{'caseName': 'Disk', 'priority': '中'}]],
        (str(tmp_path),),
    ]
    with patch('tools.export_collection_excel.get_db', _fake_db_seq(results)):
        from tools.export_collection_excel import handle
        res = handle({'collection': 'ic'}, _ctx('developer'))
    assert res['saved'] is True
    assert res['rows'] == 2
    assert res['path'].startswith('outputs/ic-')
    files = list((tmp_path / 'outputs').glob('*.xlsx'))
    assert len(files) == 1
    from openpyxl import load_workbook
    rows = list(load_workbook(files[0]).active.iter_rows(values_only=True))
    assert rows[0] == ('用例名称', '优先级')
    assert rows[1][0] == 'CPU'


def test_export_role_denied(tmp_path):
    results = [('page-ic', '保密表', ['admin'], [])]
    with patch('tools.export_collection_excel.get_db', _fake_db_seq(results)):
        from tools.export_collection_excel import handle, ExportError
        with pytest.raises(ExportError):
            handle({'collection': 'ic'}, _ctx('guest'))


def test_export_unknown_collection(tmp_path):
    results = [None]
    with patch('tools.export_collection_excel.get_db', _fake_db_seq(results)):
        from tools.export_collection_excel import handle, ExportError
        with pytest.raises(ExportError):
            handle({'collection': 'ghost'}, _ctx('admin'))


def test_export_requires_collection():
    from tools.export_collection_excel import handle, ExportError
    with pytest.raises(ExportError):
        handle({}, _ctx('admin'))


def test_export_rejects_kefu_guest():
    """kefu-guest must be rejected before any DB call."""
    from tools.export_collection_excel import handle, ExportError
    with pytest.raises(ExportError, match="not available for public customer-service sessions"):
        handle({'collection': 'any'}, _ctx('kefu-guest'))
