"""Tests for utils.data_export (intent detection + xlsx export)."""

import os
import sys
import pytest
from unittest.mock import patch
from contextlib import contextmanager

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from utils.data_export import is_export_intent, resolve_collection_from_text, export_collection_to_xlsx, ExportError


def test_is_export_intent():
    assert is_export_intent('请把巡检用例数据导出成 excel 文件')
    assert is_export_intent('导出表格')
    assert is_export_intent('生成一份数据报表')
    assert not is_export_intent('你好，今天天气如何')
    assert not is_export_intent('写一个排序算法')


class _Cur:
    def __init__(self, script):
        self._script = script
        self._result = None
    def execute(self, sql, params=None):
        for needle, rows in self._script:
            if needle in sql:
                self._result = rows
                return
        self._result = []
    def fetchall(self):
        return self._result
    def fetchone(self):
        return self._result[0] if self._result else None


def _fake_db(script):
    class _Conn:
        def cursor(self):
            return _Cur(script)
    @contextmanager
    def _get():
        yield _Conn()
    return _get


def test_resolve_collection_from_text_matches_longest_name():
    script = [("SELECT page_id, name FROM menus", [
        ('page-inspection-case', '巡检用例'),
        ('page-special-record', '专项巡检'),
    ])]
    with patch('utils.data_export.get_db', _fake_db(script)):
        res = resolve_collection_from_text('请把巡检用例数据导出成 excel')
    assert res == ('inspection-case', '巡检用例')


def test_resolve_collection_none_when_no_match():
    script = [("SELECT page_id, name FROM menus", [('page-x', '订单')])]
    with patch('utils.data_export.get_db', _fake_db(script)):
        assert resolve_collection_from_text('随便聊聊') is None


def test_export_writes_xlsx_with_field_headers(tmp_path):
    ws = str(tmp_path)
    script = [
        ("FROM menus m", [('page-ic', '巡检用例', ['admin', 'developer'])]),
        ("FROM page_configs", [[[
            {'fieldName': 'caseName', 'label': '用例名称', 'controlType': 'text'},
            {'fieldName': 'priority', 'label': '优先级', 'controlType': 'select'},
            {'fieldName': 'checkItems', 'label': '巡检项目', 'controlType': 'multiSelect'},
        ]]]),
        ("FROM dynamic_data", [
            [{'caseName': 'CPU 检查', 'priority': '高', 'checkItems': ['a', 'b']}],
            [{'caseName': '磁盘检查', 'priority': '中', 'checkItems': ['c']}],
        ]),
    ]
    with patch('utils.data_export.get_db', _fake_db(script)):
        res = export_collection_to_xlsx('ic', ws, role='developer')
    assert res['rows'] == 2
    assert res['path'].startswith('outputs/ic-')
    out = tmp_path / 'outputs'
    files = list(out.glob('*.xlsx'))
    assert len(files) == 1
    # verify content with openpyxl
    from openpyxl import load_workbook
    wb = load_workbook(files[0])
    sh = wb.active
    rows = list(sh.iter_rows(values_only=True))
    assert rows[0] == ('用例名称', '优先级', '巡检项目')
    assert rows[1][0] == 'CPU 检查'
    assert rows[1][2] == '["a", "b"]'  # list JSON-encoded


def test_export_role_denied(tmp_path):
    script = [("FROM menus m", [('page-ic', '保密表', ['admin'])])]
    with patch('utils.data_export.get_db', _fake_db(script)):
        with pytest.raises(ExportError):
            export_collection_to_xlsx('ic', str(tmp_path), role='guest')


def test_export_unknown_collection(tmp_path):
    script = [("FROM menus m", [])]
    with patch('utils.data_export.get_db', _fake_db(script)):
        with pytest.raises(ExportError):
            export_collection_to_xlsx('ghost', str(tmp_path), role='admin')
