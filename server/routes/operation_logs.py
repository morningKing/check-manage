"""
Operation log management routes.

Endpoints:
  GET    /operationLogs          - list logs with filtering & pagination
  DELETE /operationLogs/<log_id> - delete a single log
  GET    /operationLogs/export   - export filtered logs as Excel/CSV
"""
import io
import time
from flask import Blueprint, request, jsonify, send_file
from db import get_db
from auth import admin_required
from datetime import timezone

operation_logs_bp = Blueprint('operation_logs', __name__)

# 分支名称缓存（避免每次请求都 SELECT 全表 collection_versions）
_branch_name_cache = {'map': {}, 'ts': 0}
_BRANCH_CACHE_TTL = 60  # 秒

ACTION_LABELS = {'create': '新增', 'update': '修改', 'delete': '删除'}
TARGET_LABELS = {
    'menu': '菜单', 'page_config': '页面配置',
    'dynamic_data': '动态数据', 'user': '用户', 'relation': '关联关系',
}
ROLE_LABELS = {'admin': '管理员', 'developer': '开发人员', 'guest': '访客'}


def format_ts(dt):
    if dt is None:
        return None
    if hasattr(dt, 'astimezone'):
        dt = dt.astimezone(timezone.utc)
    return dt.strftime('%Y-%m-%dT%H:%M:%S.000Z')


def _get_branch_name_map(cur):
    """获取分支名称映射（带模块级缓存）"""
    now = time.time()
    if now - _branch_name_cache['ts'] < _BRANCH_CACHE_TTL and _branch_name_cache['map']:
        return _branch_name_cache['map']
    cur.execute('SELECT id, name FROM collection_versions')
    m = {row[0]: row[1] for row in cur.fetchall()}
    _branch_name_cache['map'] = m
    _branch_name_cache['ts'] = now
    return m


def row_to_dict(row, branch_name_map=None):
    raw_branch_id = row[12] if len(row) > 12 else None
    # 根据分支ID获取显示名称
    if raw_branch_id and raw_branch_id != 'main':
        branch_name = branch_name_map.get(raw_branch_id, raw_branch_id) if branch_name_map else raw_branch_id
    else:
        branch_name = '主分支'
    return {
        'id': row[0],
        'action': row[1],
        'targetType': row[2],
        'targetId': row[3],
        'targetName': row[4],
        'description': row[5],
        'operatorId': row[6],
        'operatorName': row[7],
        'operatorRole': row[8],
        'createdAt': format_ts(row[9]),
        'batchId': row[10],
        'batchDesc': row[11],
        'branchName': branch_name,
    }


COLUMNS = 'id, action, target_type, target_id, target_name, description, operator_id, operator_name, operator_role, created_at, batch_id, batch_desc, branch_id'


def build_filter():
    """Build WHERE clause and params from query string."""
    conditions = []
    params = []

    action = request.args.get('action', '').strip()
    target_type = request.args.get('targetType', '').strip()
    operator_name = request.args.get('operatorName', '').strip()
    start_time = request.args.get('startTime', '').strip()
    end_time = request.args.get('endTime', '').strip()
    branch_id = request.args.get('branchId', '').strip()

    if action:
        conditions.append('action = %s')
        params.append(action)
    if target_type:
        conditions.append('target_type = %s')
        params.append(target_type)
    if operator_name:
        conditions.append('operator_name LIKE %s')
        params.append(f'%{operator_name}%')
    if start_time:
        conditions.append('created_at >= %s')
        params.append(start_time)
    if end_time:
        conditions.append('created_at <= %s')
        params.append(end_time)
    if branch_id:
        conditions.append('branch_id = %s')
        params.append(branch_id)

    where = ''
    if conditions:
        where = 'WHERE ' + ' AND '.join(conditions)
    return where, params


@operation_logs_bp.route('/operationLogs', methods=['GET'])
@admin_required
def list_logs():
    """List operation logs with optional filtering and pagination."""
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('pageSize', 20, type=int)
    where, params = build_filter()

    with get_db() as conn:
        cur = conn.cursor()

        # 获取分支名称映射（带缓存）
        branch_name_map = _get_branch_name_map(cur)

        cur.execute(f'SELECT COUNT(*) FROM operation_logs {where}', params)
        total = cur.fetchone()[0]

        offset = (page - 1) * page_size
        cur.execute(
            f'SELECT {COLUMNS} FROM operation_logs {where} ORDER BY created_at DESC LIMIT %s OFFSET %s',
            params + [page_size, offset],
        )
        rows = cur.fetchall()

    return jsonify({
        'items': [row_to_dict(r, branch_name_map) for r in rows],
        'total': total,
    })


@operation_logs_bp.route('/operationLogs/<log_id>', methods=['DELETE'])
@admin_required
def delete_log(log_id):
    """Delete a single operation log entry."""
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute('DELETE FROM operation_logs WHERE id = %s', (log_id,))
    return jsonify({})


@operation_logs_bp.route('/operationLogs/export', methods=['GET'])
@admin_required
def export_logs():
    """Export filtered logs as an Excel or CSV file."""
    where, params = build_filter()

    with get_db() as conn:
        cur = conn.cursor()
        # 获取分支名称映射（带缓存）
        branch_name_map = _get_branch_name_map(cur)

        cur.execute(
            f'SELECT {COLUMNS} FROM operation_logs {where} ORDER BY created_at DESC',
            params,
        )
        rows = cur.fetchall()

    headers = ['操作描述', '操作类型', '目标类型', '目标名称', '操作人', '角色', '分支', '时间', '批次']

    def make_row(r):
        d = row_to_dict(r, branch_name_map)
        return [
            d['description'],
            ACTION_LABELS.get(d['action'], d['action']),
            TARGET_LABELS.get(d['targetType'], d['targetType']),
            d['targetName'] or '',
            d['operatorName'],
            ROLE_LABELS.get(d['operatorRole'], d['operatorRole']),
            d['branchName'],
            d['createdAt'] or '',
            d['batchDesc'] or '',
        ]

    # Try openpyxl for Excel output
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '操作日志'
        ws.append(headers)

        for col_num in range(1, len(headers) + 1):
            ws.cell(row=1, column=col_num).font = openpyxl.styles.Font(bold=True)

        for r in rows:
            ws.append(make_row(r))

        for col_num in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_num)
            max_length = max(
                (len(str(cell.value or '')) for cell in ws[col_letter]),
                default=10,
            )
            ws.column_dimensions[col_letter].width = min(max_length + 4, 50)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='operation_logs.xlsx',
        )
    except ImportError:
        pass

    # Fallback: CSV with UTF-8 BOM
    import csv
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    for r in rows:
        writer.writerow(make_row(r))
    output.seek(0)

    return send_file(
        io.BytesIO(output.getvalue().encode('utf-8-sig')),
        mimetype='text/csv',
        as_attachment=True,
        download_name='operation_logs.csv',
    )
